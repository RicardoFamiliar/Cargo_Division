import tkinter as tk
from tkinter import filedialog
import tkinter.messagebox as messagebox
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import customtkinter as Custom
import subprocess

class App(Custom.CTk):
    def __init__(root):
        super().__init__()

        root.imported_file_path = ""

        # Function to read the data from the Excel file
        def read_excel_data(file_path):
            try:
                global error
                error = 0
                df = pd.read_excel(file_path)
                items = [(row["EXPEDITEUR"], row["KGS"], row["M3"], row["MEAD"]) for _, row in df.iterrows()]
                
                # Get the weight and meters values from the entry widgets
                weight = weight_entry.get()
                meters = meters_entry.get()


                # Convert weight and meters to float
                try:
                    truck_capacity = {
                        'weight': float(weight),
                        'meters': float(meters),

                    }
                except ValueError as e:
                    # messagebox.showerror("Error", f"Error: {e}")
                    return [], {}

                return items, truck_capacity
            except pd.errors.ParserError as pe:
                messagebox.showerror("Error", f"Error parsing the Excel file: {pe}")
                error +=1
                return 
            except KeyError as ke:
                messagebox.showerror("Error", f"Required column not found in the Excel file: {ke}")
                error +=1
                return 


        def custom_first_fit_decreasing(items, truck_capacity):
            sorted_items = sorted(items, key=lambda x: x[1] + x[2], reverse=True)
            trucks = []

            casablanca_items = []
            tanger_items = []
            other_items = []

            for item in sorted_items:
                name, weight, meters, mead = item

                if mead == 'CASABLANCA':
                    casablanca_items.append(item)
                elif mead == 'TANGER':
                    tanger_items.append(item)
                else:
                    other_items.append(item)

            # Combine items from CASABLANCA and TANGER
            combined_items = casablanca_items + tanger_items

            for item in combined_items:
                name, weight, meters, mead = item

                # Check if there's an existing truck with enough capacity for the item
                matching_truck = next((truck for truck in trucks if truck['weight'] + weight <= truck_capacity['weight'] and truck['meters'] + meters <= truck_capacity['meters']), None)

                if matching_truck:
                    # Pack the item in the existing matching truck
                    matching_truck['items'].append(item)
                    matching_truck['weight'] += weight
                    matching_truck['meters'] += meters
                else:
                    # Create a new truck for the item
                    new_truck = {
                        'name': name,
                        'items': [item],
                        'weight': weight,
                        'meters': meters
                    }
                    if new_truck['weight'] <= truck_capacity['weight'] and new_truck['meters'] <= truck_capacity['meters']:
                        trucks.append(new_truck)
                    else:
                        # Split items that exceed the truck capacity
                        split_trucks = split_items([item], truck_capacity)
                        trucks.extend(split_trucks)

            # Pack the rest of the items using the standard first-fit-decreasing approach
            for item in other_items:
                added = False
                for truck in trucks:
                    if truck['weight'] + item[1] <= truck_capacity['weight'] and truck['meters'] + item[2] <= truck_capacity['meters']:
                        truck['items'].append(item)
                        truck['weight'] += item[1]
                        truck['meters'] += item[2]
                        added = True
                        break

                if not added:
                    new_truck = {
                        'name': f"Truck {len(trucks) + 1}",
                        'items': [item],
                        'weight': item[1],
                        'meters': item[2]
                    }
                    if new_truck['weight'] <= truck_capacity['weight'] and new_truck['meters'] <= truck_capacity['meters']:
                        trucks.append(new_truck)
                    else:
                        split_trucks = split_items([item], truck_capacity)
                        trucks.extend(split_trucks)

            return trucks

        def split_items(items, truck_capacity):
            # Helper function to split items that exceed the truck capacity
            trucks = []
            current_truck = {
                'name': f"Truck 1",
                'items': [],
                'weight': 0,
                'meters': 0
            }

            for item in items:
                name, weight, meters = item

                if current_truck['weight'] + weight <= truck_capacity['weight'] and current_truck['meters'] + meters <= truck_capacity['meters']:
                    current_truck['items'].append(item)
                    current_truck['weight'] += weight
                    current_truck['meters'] += meters
                else:
                    trucks.append(current_truck)
                    current_truck = {
                        'name': f"Truck {len(trucks) + 1}",
                        'items': [item],
                        'weight': weight,
                        'meters': meters
                    }

            if current_truck['items']:
                trucks.append(current_truck)

            return trucks

        def import_excel_file():
            file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
            if file_path:
                root.imported_file_path = file_path
                print("Excel file imported successfully.")
                file_name = file_path.split("/")[-1]  
                import_label.configure(text=f"File Imported: {file_name}")


        # Function to handle the "Execute and Export" button click event
        def execute_and_export():
            if not root.imported_file_path:
                messagebox.showerror("Error","Please import an excel file first.")
                return

            items, truck_capacity = read_excel_data(root.imported_file_path)
            if (not items or not truck_capacity) and error == 0:
                messagebox.showerror("Error", "Please, insert a KG's and M3 limit.")
                return
            
            packed_trucks = custom_first_fit_decreasing(items, truck_capacity)
            
            # Create a list to store the final rows for the DataFrame
            data_rows = []

            truck_number = 1
           # total_truck_count = len(packed_trucks)

            for truck in packed_trucks:
                total_weight = truck['weight']
                total_meters = truck['meters']
                truck_name_added = False

                for item in truck['items']:
                    name, weight, meters, mead = item
                    if not truck_name_added:
                        data_rows.append([f"Truck {truck_number}", name, weight, meters, mead])
                        truck_name_added = True
                    else:
                        data_rows.append(["", name, weight, meters, mead])


                # Add a row with total weight and total meters for the truck
                data_rows.append(["", "Total Weight:", total_weight, "Total Meters:", total_meters])

                # Increment the truck number for the next iteration
                truck_number += 1

            default_file_name = "Cargo_Division.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx")],
                initialfile=default_file_name
            )
            if file_path:
                packing_results_df = pd.DataFrame(data_rows, columns=["Truck", "Item", "Weight", "Meters", "MEAD"])
                packing_results_df.to_excel(file_path, index=False)
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active

                # Define the orange color
                orange_fill = openpyxl.styles.PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                header = openpyxl.styles.PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

                # Find the rows containing "Total Weight" and "Total Meters"
                total_weight_rows = [idx + 2 for idx, val in enumerate(data_rows) if "Total Weight:" in val]
                total_meters_rows = [idx + 2 for idx, val in enumerate(data_rows) if "Total Meters:" in val]

                # Apply orange color to the total weight and total meters rows
                for row_idx in total_weight_rows + total_meters_rows:
                    for col_idx in range(1, len(data_rows[0]) + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        cell.fill = orange_fill

                header_row = sheet[1]
                for cell in header_row:
                    cell.fill = header

                for col in sheet.columns:
                    max_length = 0
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2 
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = adjusted_width

                            # Save the workbook
                workbook.save(file_path)
                print("Excel file exported successfully.")
                messagebox.showinfo("Success", "File exported successfully.")

                subprocess.Popen(['start', '', file_path], shell=True)


        #Interface using tkinter
        Custom.set_appearance_mode("Dark") 
        Custom.set_default_color_theme("blue")  

        root.title("Truck Packing Tool")
        root.geometry("420x250")


        import_label = Custom.CTkLabel(master=root, text="")
        import_label.grid(row=14, column=0, columnspan=2, padx=0, pady=0)

        weight_label = Custom.CTkLabel(root, text="Maximum Weight Capacity (KGS):")
        weight_label.grid(row=10, column=0, padx=25, pady=(20, 5), sticky='w')

        meters_label = Custom.CTkLabel(root, text="Maximum Meters Capacity (M3):")
        meters_label.grid(row=12, column=0, padx=25, pady=(0, 5), sticky='w')

        weight_entry = tk.Entry(root, state='normal',disabledbackground='#D3D3D3')
        weight_entry.grid(row=10, column=1, padx=25, pady=(20, 5), sticky='ewns')

        meters_entry = tk.Entry(root, state='normal', disabledbackground='#D3D3D3')
        meters_entry.grid(row=12, column=1, padx=25, pady=(0, 5), sticky='ewns')
    

        button_frame2 = Custom.CTkFrame(root)
        button_frame2.grid(row=15, column=0, columnspan=2, padx=50, pady=(5,10), sticky='ew')

        import_button = Custom.CTkButton(button_frame2, text="Import", command=import_excel_file, width=20)
        import_button.pack(side=tk.LEFT, padx=0)  

        execute_button = Custom.CTkButton(button_frame2, text="Export", command=execute_and_export, width=20)
        execute_button.pack(side=tk.RIGHT, padx=0)  
        root.appearance_mode_optionemenu = Custom.CTkOptionMenu(root, values=["Theme","Light", "Dark",],
                                                                            command=root.change_appearance_mode_event)
        root.appearance_mode_optionemenu.grid(row=6, column=0, padx=(7,100), pady=(3, 0))

        root.appearance_mode_optionemenu = Custom.CTkOptionMenu(root, values=["Theme","Light", "Dark",],
                                                                            command=root.change_appearance_mode_event)
        root.appearance_mode_optionemenu.grid(row=6, column=0, padx=(7,100), pady=(3, 0))
    def change_appearance_mode_event(root, new_appearance_mode: str):
        Custom.set_appearance_mode(new_appearance_mode)


if __name__ == "__main__":
    app = App()
    app.mainloop()
